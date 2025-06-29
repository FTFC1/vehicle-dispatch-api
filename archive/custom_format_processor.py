#!/usr/bin/env python3
"""
Custom Format Report Processor
Generate Changan and Maxus reports in specific formats based on templates.
"""

import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def clean_excel_characters(text):
    """Clean invalid Excel characters from text."""
    if pd.isna(text) or not isinstance(text, str):
        return text
    
    import re
    # Clean problematic characters for Excel
    cleaned = re.sub(r'[:\\/*?\[\]]', '-', text)
    cleaned = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cleaned)
    cleaned = cleaned.replace('\'', '').replace('"', '').replace('&', 'and')
    
    if len(cleaned) > 32000:
        cleaned = cleaned[:32000]
    
    return cleaned

def create_summary_tab(df_changan, df_maxus, worksheet):
    """Create a summary overview tab with key metrics across both brands."""
    print("Creating Summary tab...")
    
    # Summary statistics
    total_changan = len(df_changan)
    total_maxus = len(df_maxus)
    total_vehicles = total_changan + total_maxus
    
    # Get unique customers
    changan_customers = df_changan['Customer Name'].nunique()
    maxus_customers = df_maxus['Customer Name'].nunique()
    
    # Summary data
    summary_data = [
        ['VEHICLE DISPATCH SUMMARY', '', ''],
        ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M'), ''],
        ['', '', ''],
        ['BRAND BREAKDOWN', '', ''],
        ['Changan Vehicles:', total_changan, ''],
        ['Maxus Vehicles:', total_maxus, ''],
        ['Total Vehicles:', total_vehicles, ''],
        ['', '', ''],
        ['CUSTOMER ANALYSIS', '', ''],
        ['Changan Customers:', changan_customers, ''],
        ['Maxus Customers:', maxus_customers, ''],
        ['', '', ''],
        ['DATA QUALITY', '', ''],
        ['Unique Changan VINs:', df_changan['VIN'].nunique(), ''],
        ['Unique Maxus VINs:', df_maxus['VIN'].nunique(), ''],
        ['Changan Duplicates Removed:', total_changan - df_changan['VIN'].nunique(), ''],
        ['', '', ''],
        ['TOP CUSTOMERS (CHANGAN)', '', ''],
    ]
    
    # Add top Changan customers
    top_changan = df_changan['Customer Name'].value_counts().head(5)
    for customer, count in top_changan.items():
        summary_data.append([customer[:30], count, 'vehicles'])
    
    summary_data.extend([
        ['', '', ''],
        ['TOP CUSTOMERS (MAXUS)', '', ''],
    ])
    
    # Add top Maxus customers
    top_maxus = df_maxus['Customer Name'].value_counts().head(5)
    for customer, count in top_maxus.items():
        summary_data.append([customer[:30], count, 'vehicles'])
    
    # Write data to worksheet
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Style headers and important rows
            if 'SUMMARY' in str(value) or 'BREAKDOWN' in str(value) or 'ANALYSIS' in str(value) or 'QUALITY' in str(value) or 'CUSTOMERS' in str(value):
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            elif row_data[0] in ['Total Vehicles:', 'Report Generated:']:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    
    # Auto-adjust column widths
    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15

def create_combined_report(df_changan, df_maxus, output_dir):
    """Create a single Excel file with Summary, Changan, and Maxus tabs."""
    print("Creating combined three-tab Excel report...")
    
    # Clean data first
    for col in df_changan.select_dtypes(include=['object']).columns:
        df_changan[col] = df_changan[col].apply(clean_excel_characters)
    
    for col in df_maxus.select_dtypes(include=['object']).columns:
        df_maxus[col] = df_maxus[col].apply(clean_excel_characters)
    
    # Remove duplicates
    changan_initial = len(df_changan)
    df_changan_clean = df_changan.drop_duplicates(subset=['VIN'], keep='first')
    changan_final = len(df_changan_clean)
    
    maxus_initial = len(df_maxus)
    df_maxus_clean = df_maxus.drop_duplicates(subset=['VIN'], keep='first')
    maxus_final = len(df_maxus_clean)
    
    if changan_initial != changan_final:
        print(f"  Removed {changan_initial - changan_final} duplicate Changan VINs")
    
    # Prepare Changan data in sales format
    changan_formatted = pd.DataFrame({
        'VIN': df_changan_clean['VIN'],
        'Retail Date': datetime.now().strftime('%Y-%m-%d'),
        'Dealer Code': df_changan_clean.get('Branch', 'N/A'),
        'Customer Name': df_changan_clean['Customer Name'],
        'Purpose': 'Dispatch',
        'Mobile': df_changan_clean.get('Del. Contact No', 'N/A'),
        'City': 'N/A',
        'Showroom': df_changan_clean.get('Desp. Warehouse', 'N/A')
    })
    
    # Prepare Maxus data in dispatch format
    maxus_formatted = pd.DataFrame({
        'S/N': range(1, len(df_maxus_clean) + 1),
        'VIN': df_maxus_clean['VIN'],
        'Engine No': df_maxus_clean['Engine'],
        'Model': df_maxus_clean.get('Item Description', 'N/A'),
        'Customer Name': df_maxus_clean['Customer Name'],
        'Delivery Date': df_maxus_clean.get('Delivery Date', datetime.now().strftime('%Y-%m-%d')),
        'Invoice No': df_maxus_clean.get('Invoice No', 'N/A'),
        'Qty': df_maxus_clean.get('Desp. Qty', 1),
        'Warehouse': df_maxus_clean.get('Desp. Warehouse', 'N/A'),
        'Contact': df_maxus_clean.get('Del. Contact No', 'N/A'),
        'Branch': df_maxus_clean.get('Branch', 'N/A'),
        'Status': 'Dispatched'
    })
    
    # Create Excel file with three tabs
    filename = f"Vehicle_Dispatch_Report_{datetime.now().strftime('%B_%Y')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Create Summary tab
        summary_df = pd.DataFrame()  # Empty df to create the sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        summary_ws = writer.sheets['Summary']
        create_summary_tab(df_changan_clean, df_maxus_clean, summary_ws)
        
        # Create Changan tab
        changan_formatted.to_excel(writer, sheet_name='Changan', index=False)
        changan_ws = writer.sheets['Changan']
        
        # Style Changan tab
        header_font = Font(bold=True, size=11)
        changan_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for cell in changan_ws[1]:
            cell.font = header_font
            cell.fill = changan_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust Changan column widths
        for i, col in enumerate(changan_formatted.columns):
            max_width = max(
                changan_formatted[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            adjusted_width = min(max_width * 1.1, 30)
            col_letter = chr(65 + i)
            changan_ws.column_dimensions[col_letter].width = adjusted_width
        
        # Create Maxus tab
        maxus_formatted.to_excel(writer, sheet_name='Maxus', index=False)
        maxus_ws = writer.sheets['Maxus']
        
        # Style Maxus tab
        maxus_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        
        for cell in maxus_ws[1]:
            cell.font = header_font
            cell.fill = maxus_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust Maxus column widths
        for i, col in enumerate(maxus_formatted.columns):
            max_width = max(
                maxus_formatted[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            adjusted_width = min(max_width * 1.1, 25)
            col_letter = chr(65 + i)
            maxus_ws.column_dimensions[col_letter].width = adjusted_width
    
    print(f"  ✓ Combined report saved: {filepath}")
    print(f"  ✓ Summary tab: Overview and key metrics")
    print(f"  ✓ Changan tab: {len(changan_formatted)} records (Sales format)")
    print(f"  ✓ Maxus tab: {len(maxus_formatted)} records (Dispatch format)")
    
    return filepath

def analyse_duplicate_issue(df_changan):
    """Analyse the duplicate issue in detail."""
    print("\n🔍 DUPLICATE ANALYSIS:")
    
    # Check for duplicate VINs
    vin_duplicates = df_changan[df_changan['VIN'].duplicated(keep=False)]
    if not vin_duplicates.empty:
        print(f"  Found {len(vin_duplicates)} duplicate VIN records:")
        for vin in vin_duplicates['VIN'].unique():
            vin_records = df_changan[df_changan['VIN'] == vin]
            print(f"    VIN {vin}: {len(vin_records)} occurrences")
            for idx, row in vin_records.iterrows():
                print(f"      - Engine: {row['Engine']}, Customer: {row['Customer Name']}")
    
    # Check for duplicate engines
    engine_duplicates = df_changan[df_changan['Engine'].duplicated(keep=False)]
    if not engine_duplicates.empty:
        print(f"  Found {len(engine_duplicates)} duplicate Engine records:")
        for engine in engine_duplicates['Engine'].unique():
            engine_records = df_changan[df_changan['Engine'] == engine]
            print(f"    Engine {engine}: {len(engine_records)} occurrences")
            for idx, row in engine_records.iterrows():
                print(f"      - VIN: {row['VIN']}, Customer: {row['Customer Name']}")

def main():
    print("\n=== CUSTOM FORMAT REPORT PROCESSOR ===\n")
    
    # Load the existing processed data
    source_file = "Files/output/Dispatch Report 06 - 2025.xlsx"
    
    if not os.path.exists(source_file):
        print(f"Error: Source file not found: {source_file}")
        print("Please run the main processor first to generate the base report.")
        return
    
    output_dir = "Files/output"
    
    try:
        # Load Changan and Maxus data
        print("Loading processed data...")
        df_changan = pd.read_excel(source_file, sheet_name='Changan')
        df_maxus = pd.read_excel(source_file, sheet_name='Maxus')
        
        print(f"  Changan: {len(df_changan)} records")
        print(f"  Maxus: {len(df_maxus)} records")
        
        # Analyse duplicate issue
        analyse_duplicate_issue(df_changan)
        
        print("\n" + "="*50)
        
        # Create combined three-tab report
        combined_file = create_combined_report(df_changan, df_maxus, output_dir)
        
        print("\n=== COMBINED REPORT PROCESSING COMPLETE ===")
        print(f"Three-tab Excel file: {os.path.basename(combined_file)}")
        print("📊 Summary | 🚗 Changan | 🚚 Maxus")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 