#!/usr/bin/env python3
"""
Vehicle Dispatch Report Processor - Non-interactive Version
Split Engine-VIN pairs into separate rows for Changan, Maxus, and Geely.
Consolidate results into a single Excel file with multiple worksheets.
"""

import pandas as pd
import os
import sys
import re
import glob
from pathlib import Path
from datetime import datetime

# Define the known column names from the raw XLS file
KNOWN_COLUMN_NAMES = [
    "Customer Name", 
    "Item Code", 
    "Item Description", 
    "Delivery No", 
    "Delivery Date", 
    "Invoice No", 
    "Invoice Date", 
    "Inv. Qty", 
    "Desp. Qty", 
    "Pend. Qty", 
    "Engine-Alternator No.", 
    "Site ID", 
    "Reference No.", 
    "Delivery AT", 
    "Del. Contact No", 
    "Desp. Warehouse", 
    "Cust.Rec.No", 
    "Cust.Rec.Date", 
    "Branch", 
    "Return Qty"
]

def clean_excel_characters(text):
    """Clean invalid Excel characters from text."""
    if pd.isna(text) or not isinstance(text, str):
        return text
    
    # More aggressive cleaning for Excel compatibility
    # Replace : / \ ? * [ ] with dashes
    cleaned = re.sub(r'[:\\/*?\[\]]', '-', text)
    
    # Replace other potential problematic characters
    cleaned = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', cleaned)  # Control characters
    cleaned = cleaned.replace('\'', '')      # Apostrophes can cause issues
    cleaned = cleaned.replace('"', '')       # Double quotes can cause issues
    cleaned = cleaned.replace('&', 'and')    # Ampersands can be problematic
    cleaned = cleaned.replace('<', '(')      # Replace angle brackets
    cleaned = cleaned.replace('>', ')')
    cleaned = cleaned.replace('#', 'No.')    # Hash can cause issues
    
    # Handle specific problematic phrases we found
    if "CHRIST EMBASSY RD" in cleaned or "LAGOS" in cleaned:
        cleaned = cleaned.replace("CHRIST EMBASSY RD", "CHRIST EMBASSY ROAD")
        cleaned = cleaned.replace("LAGOSIBADAN", "LAGOS-IBADAN")
    
    # Trim to max Excel cell length (32,767 characters)
    if len(cleaned) > 32000:
        cleaned = cleaned[:32000]
    
    return cleaned

def drop_empty_columns(df):
    """Remove empty or all-NaN columns from a dataframe."""
    # Check which columns are completely empty or all NaN
    empty_cols = []
    for col in df.columns:
        # Check if column is all NaN or empty strings
        if df[col].isna().all() or (df[col].astype(str).str.strip() == '').all():
            empty_cols.append(col)
    
    # Drop the empty columns
    if empty_cols:
        print(f"  Removing {len(empty_cols)} empty columns")
        df = df.drop(columns=empty_cols)
    
    return df

def generate_combined_report(processed_data_by_brand, engine_vin_col, output_dir):
    """Generate a single Excel file with multiple worksheets for each brand plus a summary."""
    # Use the requested naming convention: "Dispatch Report MM - YYYY"
    current_date = datetime.now()
    year_full = current_date.strftime("%Y")  # Get year as four digits
    month = current_date.strftime("%m")       # Get month as two digits
    
    combined_file = os.path.join(output_dir, f"Dispatch Report {month} - {year_full}.xlsx")
    print(f"\nGenerating combined report file: {combined_file}")
    
    # First clean up the data
    raw_data = None
    for brand_name in processed_data_by_brand:
        # Store a copy of raw data before filtering columns
        if raw_data is None:
            raw_data = processed_data_by_brand[brand_name].copy()
        else:
            raw_data = pd.concat([raw_data, processed_data_by_brand[brand_name]])
            
        # Remove empty columns and unnamed columns
        processed_data_by_brand[brand_name] = drop_empty_columns(processed_data_by_brand[brand_name])
        
        # Get column names without "Unnamed"
        columns_to_drop = [col for col in processed_data_by_brand[brand_name].columns if 'unnamed' in str(col).lower()]
        if columns_to_drop:
            print(f"  Removing {len(columns_to_drop)} 'Unnamed' columns from {brand_name}")
            processed_data_by_brand[brand_name] = processed_data_by_brand[brand_name].drop(columns=columns_to_drop)
    
    # Create Excel writer object with openpyxl
    with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
        # First create a summary tab
        summary_data = []
        total_vehicles = 0
        
        for brand_name, brand_df in processed_data_by_brand.items():
            # Check if 'VIN' column exists
            if 'VIN' not in brand_df.columns:
                print(f"  WARNING: 'VIN' column not found in {brand_name} data. Adding empty VIN column.")
                brand_df['VIN'] = ''
                processed_data_by_brand[brand_name] = brand_df  # Update the main dictionary
                
            # Check if 'Engine' column exists
            if 'Engine' not in brand_df.columns:
                print(f"  WARNING: 'Engine' column not found in {brand_name} data. Adding empty Engine column.")
                brand_df['Engine'] = ''
                processed_data_by_brand[brand_name] = brand_df  # Update the main dictionary
                
            # Count vehicles with non-empty VINs
            valid_vehicles = brand_df[~brand_df['VIN'].isna() & (brand_df['VIN'] != '')].shape[0]
            total_vehicles += valid_vehicles
            
            # Count unique engines and VINs
            unique_engines = brand_df['Engine'].nunique()
            unique_vins = brand_df['VIN'].nunique()
            
            # Add brand summary without date range
            summary_data.append({
                'Brand': brand_name,
                'Total Vehicles': valid_vehicles,
                'Unique Engines': unique_engines,
                'Unique VINs': unique_vins
            })
        
        # Create summary dataframe
        summary_df = pd.DataFrame(summary_data)
        
        # Add summary totals row
        summary_df.loc[len(summary_df)] = {
            'Brand': 'TOTAL',
            'Total Vehicles': total_vehicles,
            'Unique Engines': summary_df['Unique Engines'].sum(),
            'Unique VINs': summary_df['Unique VINs'].sum()
        }
        
        # Write summary tab (first)
        print("  Adding Summary worksheet")
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Auto-adjust column widths for summary sheet
        worksheet = writer.sheets['Summary']
        for i, col in enumerate(summary_df.columns):
            max_width = max(
                summary_df[col].astype(str).map(len).max(),
                len(col)
            ) + 2  # Add some padding
            # Excel column width is approximate, scale it for better fit
            adjusted_width = min(max_width * 1.2, 50)  # Cap at 50
            col_letter = chr(65 + i)  # A, B, C, etc.
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Format the summary tab
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Apply total row styling
        for cell in worksheet[len(summary_df) + 1]:
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
        
        # Apply borders to all data cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(summary_df)):
            for cell in row:
                cell.border = border
        
        # Add brand tabs in the specified order: Changan, Maxus, Geely
        brand_order = ["Changan", "Maxus", "Geely"]
        
        for brand_name in brand_order:
            if brand_name in processed_data_by_brand:
                brand_df = processed_data_by_brand[brand_name]
                print(f"  Adding {brand_name} worksheet with {len(brand_df)} rows")
                
                # Clean column names
                brand_df.columns = [clean_excel_characters(str(col)) for col in brand_df.columns]
                
                # Clean string data in all object columns
                for col in brand_df.select_dtypes(include=['object']).columns:
                    brand_df[col] = brand_df[col].apply(clean_excel_characters)
                
                # Select only the required columns (if they exist in the dataframe)
                required_columns = ['Customer Name', 'Item Code', 'Item Description', 'Engine', 'VIN']
                existing_required_cols = [col for col in required_columns if col in brand_df.columns]
                
                # If any required columns are missing, print a warning
                missing_cols = [col for col in required_columns if col not in existing_required_cols]
                if missing_cols:
                    print(f"  Warning: Missing columns in {brand_name}: {', '.join(missing_cols)}")
                    
                # If the dataset has the required columns, use only those
                if existing_required_cols:
                    simplified_df = brand_df[existing_required_cols].copy()
                    print(f"  Kept only {len(existing_required_cols)} required columns for {brand_name}")
                else:
                    # Fallback to using all columns
                    simplified_df = brand_df.copy()
                    print(f"  Warning: Could not find required columns, using all columns for {brand_name}")
                
                # Write to Excel with the cleaned data
                simplified_df.to_excel(writer, sheet_name=brand_name, index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets[brand_name]
                for i, col in enumerate(simplified_df.columns):
                    # Find maximum length in column
                    if col in simplified_df.select_dtypes(include=['object']).columns:
                        col_width = simplified_df[col].astype(str).map(len).max()
                    else:
                        col_width = len(str(simplified_df[col].max()))
                    
                    # Consider header width too
                    header_width = len(str(col))
                    max_width = max(col_width, header_width) + 2  # Add padding
                    
                    # Adjust width (cap at reasonable maximum to prevent very wide columns)
                    adjusted_width = min(max_width * 1.1, 30)
                    
                    # Apply the width to the column
                    col_letter = chr(65 + i) if i < 26 else chr(64 + int(i/26)) + chr(65 + (i % 26))
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # Apply styling to header row
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
        
        # Finally add the Raw data tab
        if raw_data is not None:
            print(f"  Adding RAW data tab with {len(raw_data)} rows")
            
            # Clean column names
            raw_data.columns = [clean_excel_characters(str(col)) for col in raw_data.columns]
            
            # Remove unnamed columns
            unnamed_cols = [col for col in raw_data.columns if 'unnamed' in str(col).lower()]
            if unnamed_cols:
                print(f"  Removing {len(unnamed_cols)} 'Unnamed' columns from Raw data")
                raw_data = raw_data.drop(columns=unnamed_cols)
            
            # Clean string data in all object columns
            for col in raw_data.select_dtypes(include=['object']).columns:
                raw_data[col] = raw_data[col].apply(clean_excel_characters)
            
            # Write the raw data to a separate tab
            raw_data.to_excel(writer, sheet_name='RAW', index=False)
            
            # Format the RAW tab
            worksheet = writer.sheets['RAW']
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths for raw data
            for i, col in enumerate(raw_data.columns):
                col_letter = chr(65 + i) if i < 26 else chr(64 + int(i/26)) + chr(65 + (i % 26))
                if col in raw_data.select_dtypes(include=['object']).columns:
                    col_width = raw_data[col].astype(str).map(len).max()
                else:
                    col_width = len(str(raw_data[col].max()))
                
                # Consider header width too
                header_width = len(str(col))
                max_width = max(col_width, header_width) + 2  # Add padding
                
                # Adjust width (cap at reasonable maximum)
                adjusted_width = min(max_width * 1.1, 30)
                worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✓ Successfully generated combined report: {combined_file}")
    return combined_file

def fix_column_names(df, idx_mapping):
    """
    Fix column names by using the known column names from the raw XLS file.
    
    Args:
        df: The DataFrame with auto-detected column names
        idx_mapping: A dictionary mapping column indices to their correct names
    
    Returns:
        DataFrame with fixed column names
    """
    new_columns = list(df.columns)
    
    # Replace column names based on the mapping
    for idx, correct_name in idx_mapping.items():
        if idx < len(new_columns):
            new_columns[idx] = correct_name
    
    # Update DataFrame column names
    df.columns = new_columns
    
    return df

def find_header_rows(df_raw):
    """Analyze the raw data to determine the number of header rows"""
    # Look for patterns that typically indicate headers
    potential_header_rows = []
    
    # Step 1: Look for rows with many NaN values
    for i in range(min(5, len(df_raw))):
        nan_count = df_raw.iloc[i].isna().sum()
        nan_percent = nan_count / len(df_raw.columns)
        
        if nan_percent > 0.5:  # If more than 50% of columns are empty, likely a header row
            potential_header_rows.append(i)
    
    # Step 2: Look for rows with 'customer name', 'item', 'description', etc.
    header_keywords = ['customer', 'name', 'item', 'description', 'delivery', 'invoice', 'engine']
    for i in range(min(5, len(df_raw))):
        row_values = df_raw.iloc[i].astype(str).str.lower()
        if any(keyword in ' '.join(row_values.values) for keyword in header_keywords):
            if i not in potential_header_rows:
                potential_header_rows.append(i)
    
    # Determine actual header row (last potential header row + 1)
    if potential_header_rows:
        header_row = max(potential_header_rows) + 1
    else:
        header_row = 1  # Default to row 2 (index 1) if can't determine
    
    print(f"Detected header row(s): {potential_header_rows}")
    print(f"Using row {header_row+1} as data start (after header)")
    
    return header_row, potential_header_rows

def clean_engine_number(engine_str):
    """Deep clean engine numbers to remove artifacts."""
    if not engine_str or not isinstance(engine_str, str):
        return ""
    
    # Remove trailing hyphens, dashes, asterisks
    engine_str = engine_str.rstrip('-*')
    
    # Remove asterisks in the string
    engine_str = engine_str.replace('*', '')
    
    # Handle odd formats with quotes
    engine_str = engine_str.replace('"', '')
    engine_str = engine_str.replace("'", "")
    
    # Fix spaces after hyphens (e.g., "JLH- 3G15TD" to "JLH-3G15TD")
    engine_str = re.sub(r'-\s+', '-', engine_str)
    
    # Normalize Geely engine patterns (case sensitivity)
    # Fix JL-4G15 vs Jl-4G15 (make all uppercase JL)
    engine_str = re.sub(r'^Jl-4G', 'JL-4G', engine_str)
    
    # Fix JLH-3G15TD case consistency
    if re.match(r'^JLH-3G15TD', engine_str, re.IGNORECASE):
        engine_str = re.sub(r'^[Jj][Ll][Hh]-3[Gg]15[Tt][Dd]', 'JLH-3G15TD', engine_str)
    
    # Remove all other non-alphanumeric chars except hyphen
    engine_str = re.sub(r'[^\w\-]', '', engine_str)
    
    return engine_str.strip()

def process_engine_vin_cell(raw_value):
    """
    Splits a raw cell value into multiple (engine, vin) tuples.
    Handles comma-separated values and various formatting issues.
    Returns a list of (engine, vin) tuples.
    """
    if pd.isna(raw_value) or not isinstance(raw_value, str) or not raw_value.strip():
        return [("", "")]

    # Normalize separators: replace semicolons, newlines, and spaces with commas
    normalized_value = re.sub(r'[;\n\s]+', ',', raw_value)
    engine_vin_pairs = [pair.strip() for pair in normalized_value.split(',') if pair.strip()]
    results = []

    for pair in engine_vin_pairs:
        engine, vin = "", ""
        
        # General pattern: Engine-VIN or VIN-Engine
        match = re.search(r'([A-Z0-9]+)[-]([A-Z0-9]+)', pair.upper())
        if match:
            part1, part2 = match.groups()
            # Heuristic: VIN is usually longer (17 chars)
            if len(part1) > len(part2) and len(part1) > 10:
                vin, engine = part1, part2
            else:
                engine, vin = part1, part2
        else:
            # Fallback if no hyphen is found
            vin = pair # Assume the whole thing is a VIN

        results.append((engine.strip(), vin.strip()))
        
    # If no pairs were successfully parsed, return a single empty entry
    # This ensures the explode function in the caller still works.
    if not results:
        return [("", "")]
        
    return results

def process_brands(df, engine_vin_col, brand_col, target_brands):
    """
    Filter the dataframe by brand, process the engine-VIN column, and split into separate rows.
    """
    processed_data_by_brand = {}

    for brand_name, search_terms in target_brands.items():
        # Filter by brand
        brand_query = '|'.join(search_terms)
        brand_df = df[df[brand_col].str.lower().str.contains(brand_query, na=False, regex=True)].copy()

        if not brand_df.empty:
            # 1. Apply the function to get lists of (engine, vin) tuples
            split_values = brand_df[engine_vin_col].apply(process_engine_vin_cell)

            # 2. Create an intermediate DataFrame from the lists of tuples
            split_df = pd.DataFrame(split_values.tolist(), index=brand_df.index).stack().reset_index(level=1, drop=True)
            split_df.name = 'engine_vin_pair'
            
            # 3. Merge this back to the original brand_df
            processed_df = brand_df.drop(columns=[engine_vin_col]).join(split_df)

            # 4. Separate the tuple into 'Engine' and 'VIN' columns
            processed_df[['Engine', 'VIN']] = pd.DataFrame(processed_df['engine_vin_pair'].tolist(), index=processed_df.index)

            # 5. Clean up and store
            processed_df = processed_df.drop(columns=['engine_vin_pair'])
            
            # Remove rows where VIN is empty, as they are not useful
            processed_df = processed_df[processed_df['VIN'].str.strip() != '']

            processed_data_by_brand[brand_name] = processed_df
            print(f"  Found {len(brand_df)} rows matching '{brand_query.lower()}'")
            print(f"  Created {len(processed_df)} rows after splitting")
            
    return processed_data_by_brand

def main():
    # Print header
    print("\n=== VEHICLE DISPATCH REPORT PROCESSOR - AUTOMATED VERSION ===\n")
    
    # Auto-detect all XLS files in Files directory
    source_file_path = "Files/Desp_regENDMAY2025.xls"
    print(f"Processing file: {source_file_path}")
    
    # Brands to filter for (with potential variations)
    target_brands = {
        "Changan": ["changan", "chang'an"],
        "Maxus": ["maxus", "max's", "maxs"],
        "Geely": ["geely", "gely"]
    }
    print(f"Target brands to process: {', '.join(target_brands.keys())}")
    
    # Create output directory
    output_dir = "Files/output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")
    
    # Analyze the file structure first - this is critical
    try:
        print("Analyzing Excel file structure...")
        # First try with no header to see the raw structure
        df_raw = pd.read_excel(source_file_path, sheet_name='Desp_reg', header=None, nrows=15, engine='xlrd')
        print(f"File loaded with {df_raw.shape[0]} rows and {df_raw.shape[1]} columns (raw view)")
        
        # Print the first 5 rows of raw data to understand structure
        print("\nRaw data preview (first 5 rows, first 8 columns):")
        pd.set_option('display.max_columns', 8)  # Show first 8 columns
        pd.set_option('display.max_colwidth', 25)  # Limit column width
        print(df_raw.iloc[:5, :8])  # First 5 rows, first 8 columns
        
        # Determine header row
        header_row, potential_header_rows = find_header_rows(df_raw)
        
        # Now, try with the proper header row
        df = pd.read_excel(source_file_path, sheet_name='Desp_reg', header=header_row, engine='xlrd')
        print(f"\nFile loaded with {df.shape[0]} rows and {df.shape[1]} columns (using row {header_row+1} as header)")
        
        # Create a mapping to fix column names - look for the "real" header row in potential_header_rows
        column_name_mapping = {}
        header_row_found = False
        
        for row_idx in potential_header_rows:
            row_values = df_raw.iloc[row_idx].astype(str)
            
            # Check if this row has many matches with our known column names
            matches = 0
            for known_name in KNOWN_COLUMN_NAMES:
                for i, val in enumerate(row_values):
                    if pd.notna(val) and known_name.lower() in val.lower():
                        matches += 1
                        column_name_mapping[i] = known_name
            
            # If we found multiple matches, this is likely the header row with the real column names
            if matches >= 3:  # Arbitrary threshold - at least 3 column names match
                header_row_found = True
                print(f"Found header row at index {row_idx} with {matches} matching column names")
                break
        
        # If we couldn't find a good match, use an empty mapping
        if not header_row_found:
            print("Could not find strong match for header row, using detected column names")
            column_name_mapping = {}
        
        # Print the full column names to help identify the right ones
        print("\nColumn names from detected header row:")
        for i, col in enumerate(df.columns):
            mapped_name = column_name_mapping.get(i, col)
            print(f"  Column {i+1}: {col} → {mapped_name}")
        
        # Apply the column name fixes
        if column_name_mapping:
            df = fix_column_names(df, column_name_mapping)
        
        # Find the Engine-VIN column (containing ENGINE and VIN info)
        print("\nSearching for Engine-VIN column...")
        
        # Check each column for ENGINE-VIN pattern
        engine_vin_col = None
        for i, col in enumerate(df.columns):
            # Get a sample of data from this column
            sample_data = df[col].astype(str).str.strip().dropna().head(25)
            sample_data = [s for s in sample_data if len(s) > 3 and s.lower() != 'nan']  # Filter empty values
            
            if sample_data:
                print(f"\nColumn {i+1}: {col}")
                print(f"  First few values:")
                for j, sample in enumerate(sample_data[:5]):
                    print(f"    {j+1}) {sample}")
                
                # Check if this column has the patterns we're looking for
                hyphen_count = sum(1 for s in sample_data if '-' in s)
                long_values = sum(1 for s in sample_data if len(s) > 20)
                has_comma = any(',' in s for s in sample_data)
                
                # Look for the specific column name we know contains Engine-VIN data
                if col == "Engine-Alternator No.":
                    print(f"  ✓ Found Engine-Alternator column by exact name match!")
                    engine_vin_col = col
                    print(f"  ✓ SELECTED as ENGINE-VIN column (by name)")
                    break
                
                # Look for column with 'Engine-Alternator' in header or first row
                engine_alternator_match = False
                if 'engine' in str(col).lower() and ('alternator' in str(col).lower() or 'no' in str(col).lower()):
                    engine_alternator_match = True
                elif sample_data and 'engine' in str(sample_data[0]).lower() and 'alternator' in str(sample_data[0]).lower():
                    engine_alternator_match = True
                
                if engine_alternator_match:
                    print(f"  ✓ Found Engine-Alternator column header!")
                    engine_vin_col = col
                    print(f"  ✓ SELECTED as ENGINE-VIN column (by header name)")
                    break
                
                # Also detect based on pattern
                if hyphen_count > 0 and long_values > 0:
                    print(f"  ✓ Potential ENGINE-VIN column! Found {hyphen_count}/{len(sample_data)} values with hyphens")
                    print(f"  ✓ Contains {long_values}/{len(sample_data)} long values")
                    if has_comma:
                        print(f"  ✓ Contains values with commas (multiple ENGINE-VIN pairs)")
                    
                    # This is likely our Engine-VIN column if we haven't found one by header yet
                    if not engine_vin_col and i == 10:  # Column 11 (Index 10) is the Engine-Alternator column in our file
                        engine_vin_col = col
                        print(f"  ✓ SELECTED as ENGINE-VIN column")
        
        if not engine_vin_col:
            print("\nERROR: Could not automatically find ENGINE-VIN column.")
            # Try columns 10-12 (Unnamed: 10) as fallback
            for col_idx in range(10, min(13, len(df.columns))):
                if len(df.columns) > col_idx:
                    engine_vin_col = df.columns[col_idx]
                    print(f"Attempting to use column {col_idx+1} as fallback: {engine_vin_col}")
                    break
            
            if not engine_vin_col:
                print("Cannot proceed without identifying ENGINE-VIN column.")
                return
        
        # Now find a column that has brand information (Item Description)
        print("\nSearching for Brand column...")
        brand_col = None
        
        # First, look for specific column name we know contains brand info
        for i, col in enumerate(df.columns):
            if col == "Item Description":
                brand_col = col
                print(f"Found brand column by exact name match: {col}")
                break
        
        # If not found by name, look for columns with 'customer' or 'item' in the header
        if not brand_col:
            for i, col in enumerate(df.columns):
                col_str = str(col).lower()
                if ('customer' in col_str and 'name' in col_str) or 'brand' in col_str or 'item description' in col_str:
                    brand_col = col
                    print(f"Found likely brand column by name: {col}")
                    break
        
        # If not found by name, try columns 1-3 as these often contain customer info
        if not brand_col:
            for i in range(min(3, len(df.columns))):
                col = df.columns[i]
                # Get samples to analyze
                samples = df[col].astype(str).dropna().head(20).tolist()
                samples = [s for s in samples if len(s) > 3 and s.lower() != 'nan']  # Filter empty values
                
                if samples:
                    print(f"\nAnalyzing column {i+1}: {col} for brand info:")
                    for j, sample in enumerate(samples[:5]):
                        print(f"  Sample {j+1}: {sample}")
                    
                    # Look for any brand keywords in the samples
                    found_brands = []
                    for brand, variations in target_brands.items():
                        for variation in variations:
                            matches = sum(1 for s in samples if variation in s.lower())
                            if matches > 0:
                                found_brands.append(f"{brand} ({matches} matches)")
                    
                    if found_brands:
                        print(f"  ✓ Found brand mentions: {', '.join(found_brands)}")
                        brand_col = col
                        print(f"  ✓ SELECTED as Brand column")
                        break
        
        # If still not found, check Item Description (column 3) if it exists
        if not brand_col and "Item Description" in df.columns:
            brand_col = "Item Description"
            print(f"Using 'Item Description' as Brand column")
        elif not brand_col and len(df.columns) > 2:
            brand_col = df.columns[2]  # Use column 3 as it may contain brand info
            print(f"Using column 3 as Brand column: {brand_col}")
        
        # At this point we have our best guess at engine_vin_col and brand_col
        print(f"\nFinal column selections:")
        print(f"  ENGINE-VIN column: {engine_vin_col}")
        print(f"  Brand column: {brand_col or 'None - will process all rows'}")
        
        # Since the brand detection might be challenging, let's adjust our approach
        # Instead of filtering by exact brand name, we'll create a report for each 
        # vehicle brand with ANY matching entries, even partial matches
        
        processed_data_by_brand = process_brands(df, engine_vin_col, brand_col, target_brands)
        
        # Generate the combined Excel file with multiple worksheets
        combined_file = generate_combined_report(processed_data_by_brand, engine_vin_col, output_dir)
        
        # Clean up individual files from previous runs
        print("\nCleaning up individual report files...")
        individual_files = glob.glob(os.path.join(output_dir, "*_Report*.xlsx"))
        for file in individual_files:
            if os.path.basename(file) != os.path.basename(combined_file):
                try:
                    os.remove(file)
                    print(f"  Removed {os.path.basename(file)}")
                except Exception as e:
                    print(f"  Could not remove {os.path.basename(file)}: {e}")
        
        print("\n=== PROCESSING COMPLETE ===")
        print(f"Final report can be found at: {os.path.abspath(combined_file)}")
        
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()
        return

if __name__ == "__main__":
    main() 